"""
Admin Routes for SEMIRA FASHION

All admin panel routes: dashboard, products, ads, orders, users,
inbox, reports, settings, reviews, translations, notifications.
"""

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, session, jsonify, make_response, send_from_directory, abort, current_app
)
from middleware.auth import admin_required
from database.db import get_db
from werkzeug.security import generate_password_hash, check_password_hash
from routes.shared import get_lang, WHATSAPP_NUMBER
from extensions import limiter
import os
import json
import uuid
import csv
from io import StringIO, BytesIO
import datetime as datetime_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from services.notification_service import (
    notify_user, notify_admin,
    get_admin_alerts, get_admin_unread_count, mark_admin_alerts_read
)

admin_bp = Blueprint('admin', __name__)


# ==================== ADMIN LOGIN ====================

@admin_bp.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute; 30 per hour")
def admin_login():
    if session.get('admin'):
        return redirect(url_for('admin.dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        admin_username = os.environ.get('ADMIN_USERNAME', '')
        admin_password = os.environ.get('ADMIN_PASSWORD', '')

        if not admin_username or not admin_password:
            flash('Admin login is not configured. Set ADMIN_USERNAME and ADMIN_PASSWORD in environment variables.', 'danger')
            return render_template('admin/login.html')

        # Check settings table for a stored password hash override
        _pw_ok = False
        try:
            _conn = get_db()
            _cur = _conn.cursor()
            _cur.execute("SELECT value FROM settings WHERE key = 'admin_password_hash'")
            _row = _cur.fetchone()
            if _row and _row[0]:
                _pw_ok = (username == admin_username and check_password_hash(_row[0], password))
            else:
                _pw_ok = (username == admin_username and password == admin_password)
        except Exception:
            _pw_ok = (username == admin_username and password == admin_password)

        if _pw_ok:
            session['admin'] = True
            session['admin_username'] = username
            flash('Logged in successfully!', 'success')
            next_page = request.args.get('next', '')
            from werkzeug.security import safe_join  # noqa – only for import check
            from urllib.parse import urlparse
            _p = urlparse(next_page)
            # Allow only relative URLs with no scheme/netloc (prevents open-redirect)
            if next_page and not _p.scheme and not _p.netloc and next_page.startswith('/'):
                return redirect(next_page)
            return redirect(url_for('admin.dashboard'))
        else:
            flash('Invalid username or password!', 'danger')

    return render_template('admin/login.html')


@admin_bp.route('/logout')
def admin_logout():
    session.pop('admin', None)
    session.pop('admin_username', None)
    flash('Logged out successfully!', 'success')
    return redirect(url_for('admin.admin_login'))


# ==================== ADMIN COUNTS API ====================

@admin_bp.route('/api/counts')
@admin_required
def admin_counts():
    """Return sidebar badge counts as JSON for client-side updates."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                (SELECT COUNT(*) FROM products WHERE is_active = 1),
                (SELECT COUNT(*) FROM advertisements WHERE is_active = 1),
                (SELECT COUNT(*) FROM orders WHERE status = 'pending')
        """)
        row = cursor.fetchone()
        return jsonify({
            'products': row[0] or 0,
            'ads': row[1] or 0,
            'pending_orders': row[2] or 0,
        })
    except Exception as e:
        current_app.logger.error(f"admin_counts error: {e}")
        return jsonify({'products': 0, 'ads': 0, 'pending_orders': 0})


# ==================== DASHBOARD ====================

@admin_bp.route('/')
@admin_bp.route('/dashboard')
@admin_required
def dashboard():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Single round-trip instead of 7 separate queries
        cursor.execute("""
            SELECT
                (SELECT COUNT(*) FROM products WHERE is_active = 1),
                (SELECT COUNT(*) FROM advertisements WHERE is_active = 1),
                (SELECT COUNT(*) FROM orders),
                (SELECT COUNT(*) FROM orders WHERE status = 'pending'),
                (SELECT COUNT(*) FROM users WHERE is_admin = 0),
                (SELECT COALESCE(SUM(total), 0) FROM orders WHERE status != 'cancelled'),
                (SELECT COUNT(*) FROM orders WHERE DATE(created_at) = CURRENT_DATE)
        """)
        _row = cursor.fetchone()
        products_count  = _row[0] or 0
        ads_count       = _row[1] or 0
        total_orders    = _row[2] or 0
        pending_orders  = _row[3] or 0
        customers_count = _row[4] or 0
        total_revenue   = _row[5] or 0
        today_orders    = _row[6] or 0

        cursor.execute("""
            SELECT o.*, u.full_name as customer_name
            FROM orders o LEFT JOIN users u ON o.user_id = u.id
            ORDER BY o.id DESC LIMIT 10
        """)
        recent_orders_rows = cursor.fetchall()
        recent_orders = [dict(o) for o in recent_orders_rows] if recent_orders_rows else []

        # Recent products (last 5) with category name
        cursor.execute("""
            SELECT p.id, p.name, p.name_am, p.price, p.compare_price,
                   p.stock_quantity, p.thumbnail, p.is_active,
                   c.name as category_name
            FROM products p LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.is_active = 1
            ORDER BY p.id DESC LIMIT 5
        """)
        recent_products_rows = cursor.fetchall()
        recent_products = [dict(p) for p in recent_products_rows] if recent_products_rows else []

        # Low-stock products with category name
        cursor.execute("""
            SELECT p.*, c.name as category_name
            FROM products p LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.stock_quantity <= p.low_stock_threshold AND p.stock_quantity > 0
            LIMIT 10
        """)
        low_stock_rows = cursor.fetchall()
        low_stock_products = [dict(p) for p in low_stock_rows] if low_stock_rows else []

        # Out-of-stock products with category name
        cursor.execute("""
            SELECT p.*, c.name as category_name
            FROM products p LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.stock_quantity = 0 AND p.is_active = 1
            LIMIT 10
        """)
        out_of_stock_rows = cursor.fetchall()
        out_of_stock_products = [dict(p) for p in out_of_stock_rows] if out_of_stock_rows else []

        # Page views count (graceful if table does not exist)
        try:
            cursor.execute("SELECT COUNT(*) FROM page_views")
            total_page_views = cursor.fetchone()[0] or 0
        except Exception:
            conn.rollback()
            total_page_views = 0

        stats = {
            'products_count': products_count,
            'ads_count': ads_count,
            'total_orders': total_orders,
            'pending_orders': pending_orders,
            'customers_count': customers_count,
            'total_revenue': total_revenue,
            'today_orders': today_orders,
            'live_visitors': 0,
            'total_page_views': total_page_views,
        }

        return render_template('admin/dashboard.html',
                               stats=stats,
                               recent_orders=recent_orders,
                               recent_products=recent_products,
                               low_stock_products=low_stock_products,
                               out_of_stock_products=out_of_stock_products,
                               order_stock_alerts=[],
                               lang=lang)
    except Exception as e:
        current_app.logger.error(f"Dashboard error: {e}")
        flash('Error loading dashboard.', 'error')
        return render_template('admin/dashboard.html',
                               stats={}, recent_orders=[], recent_products=[],
                               low_stock_products=[], out_of_stock_products=[],
                               order_stock_alerts=[], lang=lang)


# ==================== PRODUCT MANAGEMENT ====================

@admin_bp.route('/products')
@admin_required
def products():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.*, c.name as category_name
            FROM products p LEFT JOIN categories c ON p.category_id = c.id
            ORDER BY p.id DESC
        """)
        products_rows = cursor.fetchall()
        cursor.execute("SELECT id, name, name_am FROM categories WHERE is_active = 1 ORDER BY sort_order")
        categories_rows = cursor.fetchall()
        return render_template('admin/products/index.html',
                               products=[dict(p) for p in products_rows] if products_rows else [],
                               categories=[dict(c) for c in categories_rows] if categories_rows else [],
                               lang=lang)
    except Exception as e:
        current_app.logger.error(f"Admin products error: {e}")
        flash('Error loading products.', 'error')
        return redirect(url_for('admin.dashboard'))


@admin_bp.route('/products/create', methods=['GET', 'POST'])
@admin_required
def product_create():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, name_am FROM categories ORDER BY sort_order")
        categories = [dict(c) for c in cursor.fetchall()]

        if request.method == 'POST':
            name = request.form.get('name_en', '') or request.form.get('name', '')
            name_am = request.form.get('name_am', '')
            name_ar = request.form.get('name_ar', '')
            name_en = request.form.get('name_en', '') or name
            description = request.form.get('description_en', '') or request.form.get('description', '')
            description_am = request.form.get('description_am', '')
            description_ar = request.form.get('description_ar', '')
            description_en = request.form.get('description_en', '') or description
            try:
                price = float(request.form.get('price', 0) or 0)
                compare_price = request.form.get('compare_price')
                compare_price = float(compare_price) if compare_price else None
                stock_quantity = int(request.form.get('stock_quantity', 0) or 0)
                category_id = int(request.form.get('category_id', 0) or 0) or None
            except (ValueError, TypeError) as e:
                flash(f'Invalid number in Price, Compare Price, Stock, or Category field: {e}', 'error')
                return redirect(url_for('admin.product_create'))
            material = request.form.get('material', '')
            color = request.form.get('color', '')
            sku = request.form.get('sku', '')
            gender = request.form.get('gender', '')
            season = request.form.get('season', '')
            import json as _json
            sizes_list = request.form.getlist('sizes')
            sizes = _json.dumps(sizes_list) if sizes_list else None
            is_featured = 1 if request.form.get('is_featured') else 0
            is_new = 1 if request.form.get('is_new') else 0

            from werkzeug.utils import secure_filename
            upload_dir = 'static/uploads/products'
            os.makedirs(upload_dir, exist_ok=True)

            _MAX_IMG_BYTES = 5 * 1024 * 1024  # 5 MB per image

            _ALLOWED_IMG_EXT = {'jpg', 'jpeg', 'png', 'webp', 'gif'}

            def save_upload(file_obj):
                file_obj.seek(0, os.SEEK_END)
                size = file_obj.tell()
                file_obj.seek(0)
                if size > _MAX_IMG_BYTES:
                    raise ValueError(f"'{file_obj.filename}' exceeds 5 MB limit")
                fname = secure_filename(file_obj.filename)
                ext = fname.rsplit('.', 1)[1].lower() if '.' in fname else ''
                if ext not in _ALLOWED_IMG_EXT:
                    raise ValueError(f"File type '.{ext}' is not allowed. Use jpg, jpeg, png, webp or gif.")
                unique = f"product_{uuid.uuid4().hex[:8]}.{ext}"
                file_obj.save(os.path.join(upload_dir, unique))
                return f'uploads/products/{unique}'

            image_filename = ''
            image = request.files.get('image')
            if image and image.filename:
                image_filename = save_upload(image)

            # Handle multiple extra gallery images
            extra_images = request.files.getlist('extra_images')
            gallery_paths = []
            for extra in extra_images:
                if extra and extra.filename:
                    gallery_paths.append(save_upload(extra))

            images_json = _json.dumps(gallery_paths) if gallery_paths else None

            cursor.execute("""
                INSERT INTO products (
                    name, name_am, name_ar, name_en,
                    description, description_am, description_ar, description_en,
                    price, compare_price, stock_quantity, category_id,
                    material, color, sku, gender, season, sizes,
                    is_featured, is_new, thumbnail, images, is_active
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
            """, (name, name_am, name_ar, name_en,
                  description, description_am, description_ar, description_en,
                  price, compare_price, stock_quantity, category_id,
                  material, color, sku, gender, season, sizes,
                  is_featured, is_new, image_filename, images_json))
            conn.commit()
            flash('Product created successfully!', 'success')
            return redirect(url_for('admin.products'))

        return render_template('admin/products/create.html', categories=categories, lang=lang)
    except Exception as e:
        current_app.logger.error(f"Product create error: {e}")
        flash(f'Error creating product: {e}', 'error')
        return redirect(url_for('admin.products'))


@admin_bp.route('/products/edit/<int:pid>', methods=['GET', 'POST'])
@admin_required
def product_edit(pid):
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM products WHERE id = %s", (pid,))
        product = cursor.fetchone()
        if not product:
            flash('Product not found!', 'danger')
            return redirect(url_for('admin.products'))

        cursor.execute("SELECT id, name, name_am FROM categories ORDER BY sort_order")
        categories = [dict(c) for c in cursor.fetchall()]

        if request.method == 'POST':
            name = request.form.get('name_en', '') or request.form.get('name', '')
            name_am = request.form.get('name_am', '')
            name_ar = request.form.get('name_ar', '')
            name_en = request.form.get('name_en', '') or name
            description = request.form.get('description_en', '') or request.form.get('description', '')
            description_am = request.form.get('description_am', '')
            description_ar = request.form.get('description_ar', '')
            description_en = request.form.get('description_en', '') or description
            try:
                price = float(request.form.get('price', 0) or 0)
                compare_price = request.form.get('compare_price')
                compare_price = float(compare_price) if compare_price else None
                stock_quantity = int(float(request.form.get('stock_quantity', 0) or 0))
                category_id = int(float(request.form.get('category_id', 0) or 0)) or None
            except (ValueError, TypeError) as e:
                flash(f'Invalid number in Price, Compare Price, Stock, or Category field: {e}', 'error')
                return redirect(url_for('admin.product_edit', pid=pid))
            material = request.form.get('material', '')
            color = request.form.get('color', '')
            sku = request.form.get('sku', '')
            gender = request.form.get('gender', '')
            season = request.form.get('season', '')
            import json as _json
            sizes_list = request.form.getlist('sizes')
            sizes = _json.dumps(sizes_list) if sizes_list else None
            is_featured = 1 if request.form.get('is_featured') else 0
            is_new = 1 if request.form.get('is_new') else 0

            from werkzeug.utils import secure_filename
            upload_dir = 'static/uploads/products'
            os.makedirs(upload_dir, exist_ok=True)

            _MAX_IMG_BYTES = 5 * 1024 * 1024  # 5 MB per image
            _ALLOWED_IMG_EXT = {'jpg', 'jpeg', 'png', 'webp', 'gif'}

            def save_upload(file_obj):
                file_obj.seek(0, os.SEEK_END)
                size = file_obj.tell()
                file_obj.seek(0)
                if size > _MAX_IMG_BYTES:
                    raise ValueError(f"'{file_obj.filename}' exceeds 5 MB limit")
                fname = secure_filename(file_obj.filename)
                ext = fname.rsplit('.', 1)[1].lower() if '.' in fname else ''
                if ext not in _ALLOWED_IMG_EXT:
                    raise ValueError(f"File type '.{ext}' is not allowed. Use jpg, jpeg, png, webp or gif.")
                unique = f"product_{uuid.uuid4().hex[:8]}.{ext}"
                file_obj.save(os.path.join(upload_dir, unique))
                return f'uploads/products/{unique}'

            def _normalize_img(path):
                """Normalise a stored image path to 'uploads/products/<file>'."""
                if not path:
                    return None
                path = path.strip()
                if path.startswith('static/'):
                    path = path[len('static/'):]
                if '/' not in path:
                    path = f'uploads/products/{path}'
                return path

            image_filename = _normalize_img(product['thumbnail']) or ''
            image = request.files.get('image')
            if image and image.filename:
                image_filename = save_upload(image)

            # Existing gallery images the user kept — normalise paths from legacy data
            keep_images_raw = request.form.get('keep_images', '')
            try:
                kept_raw = _json.loads(keep_images_raw) if keep_images_raw else []
                if not isinstance(kept_raw, list):
                    kept_raw = []
                kept = [_normalize_img(p) for p in kept_raw if p]
            except Exception:
                kept = []

            # New extra gallery images
            extra_images = request.files.getlist('extra_images')
            new_gallery = []
            for extra in extra_images:
                if extra and extra.filename:
                    new_gallery.append(save_upload(extra))

            all_gallery = kept + new_gallery
            images_json = _json.dumps(all_gallery) if all_gallery else None

            # Fetch old price before update (for price-drop notifications)
            cursor.execute("SELECT price, name_am, name FROM products WHERE id = %s", (pid,))
            old_prod = cursor.fetchone()
            old_price = float(old_prod['price']) if old_prod else None

            meta_title = request.form.get('meta_title', '').strip()
            meta_description = request.form.get('meta_description', '').strip()

            cursor.execute("""
                UPDATE products SET
                    name=%s, name_am=%s, name_ar=%s, name_en=%s,
                    description=%s, description_am=%s, description_ar=%s, description_en=%s,
                    price=%s, compare_price=%s, stock_quantity=%s, category_id=%s,
                    material=%s, color=%s, sku=%s, gender=%s, season=%s, sizes=%s,
                    is_featured=%s, is_new=%s, thumbnail=%s, images=%s,
                    meta_title=%s, meta_description=%s,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id=%s
            """, (name, name_am, name_ar, name_en,
                  description, description_am, description_ar, description_en,
                  price, compare_price, stock_quantity, category_id,
                  material, color, sku, gender, season, sizes,
                  is_featured, is_new, image_filename, images_json,
                  meta_title, meta_description, pid))
            conn.commit()

            # Notify wishlist users if price dropped
            if old_price is not None and price < old_price:
                try:
                    from services.notification_service import notify_user
                    pname = (old_prod['name_am'] or old_prod['name'] or 'Product')
                    drop_pct = round((old_price - price) / old_price * 100)
                    cursor.execute(
                        "SELECT user_id FROM wishlist WHERE product_id = %s", (pid,)
                    )
                    wl_users = cursor.fetchall()
                    for wu in (wl_users or []):
                        notify_user(
                            wu['user_id'],
                            f'🔥 Price Drop on {pname}!',
                            f'Price dropped {drop_pct}% — now {price:,.0f} ETB (was {old_price:,.0f} ETB). Grab it before it\'s gone!',
                            type='price_drop',
                            link=f'/product/{pid}'
                        )
                except Exception as _ne:
                    current_app.logger.warning(f"[price-drop notify] {_ne}")

            flash('Product updated successfully!', 'success')
            return redirect(url_for('admin.products'))

        return render_template('admin/products/edit.html',
                               product=dict(product), categories=categories, lang=lang)
    except Exception as e:
        current_app.logger.error(f"Product edit error: {e}")
        flash(f'Error editing product: {e}', 'error')
        return redirect(url_for('admin.products'))


@admin_bp.route('/products/delete/<int:pid>', methods=['POST', 'DELETE'])
@admin_required
def product_delete(pid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("UPDATE products SET is_active = 0 WHERE id = %s", (pid,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Product deleted successfully'})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        current_app.logger.error("Product delete error: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Could not delete product'}), 500


@admin_bp.route('/products/duplicate/<int:pid>', methods=['GET', 'POST'])
@admin_required
def product_duplicate(pid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM products WHERE id = %s", (pid,))
        product = cursor.fetchone()
        if not product:
            return jsonify({'success': False, 'error': 'Product not found'}), 404
        p = dict(product)
        new_name = f"Copy of {p.get('name', '')}"
        new_name_am = f"ቅጂ - {p.get('name_am', '')}" if p.get('name_am') else new_name

        # Generate a guaranteed-unique SKU by appending a short random suffix
        import random, string
        base_sku = p.get('sku') or str(pid)
        while True:
            suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
            new_sku = f"COPY-{base_sku}-{suffix}"
            cursor.execute("SELECT id FROM products WHERE sku = %s", (new_sku,))
            if not cursor.fetchone():
                break

        cursor.execute("""
            INSERT INTO products (
                name, name_am, name_ar, name_en,
                description, description_am, description_ar, description_en,
                price, compare_price, stock_quantity, category_id,
                material, color, sku, gender, season, sizes,
                is_featured, is_new, thumbnail, is_active
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
            RETURNING id
        """, (
            new_name, new_name_am, p.get('name_ar'), p.get('name_en'),
            p.get('description'), p.get('description_am'), p.get('description_ar'), p.get('description_en'),
            p.get('price'), p.get('compare_price'), p.get('stock_quantity'), p.get('category_id'),
            p.get('material'), p.get('color'), new_sku,
            p.get('gender'), p.get('season'), p.get('sizes'),
            p.get('is_featured', 0), p.get('is_new', 0), p.get('thumbnail')
        ))
        row = cursor.fetchone()
        new_id = row[0] if row else None
        conn.commit()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Product duplicated successfully', 'new_id': new_id})
        flash('Product duplicated successfully!', 'success')
        return redirect(url_for('admin.products'))
    except Exception as e:
        current_app.logger.error(f"Product duplicate error: {e}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': str(e)}), 500
        flash('Error duplicating product.', 'error')
        return redirect(url_for('admin.products'))


@admin_bp.route('/products/toggle-featured/<int:pid>', methods=['POST'])
@admin_required
def product_toggle_featured(pid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT is_featured FROM products WHERE id = %s", (pid,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Product not found'}), 404
        new_state = 0 if row[0] else 1
        cursor.execute("UPDATE products SET is_featured = %s WHERE id = %s", (new_state, pid))
        conn.commit()
        return jsonify({'success': True, 'is_featured': bool(new_state)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/products/update-stock/<int:pid>', methods=['POST'])
@admin_required
def product_update_stock(pid):
    try:
        data = request.get_json()
        stock_quantity = int(data.get('stock_quantity', 0))
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("UPDATE products SET stock_quantity = %s WHERE id = %s", (stock_quantity, pid))
        conn.commit()
        return jsonify({'success': True, 'stock_quantity': stock_quantity})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/products/bulk-delete', methods=['POST'])
@admin_required
def bulk_delete_products():
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'success': False, 'error': 'No products selected'}), 400

        conn = get_db()
        cursor = conn.cursor()
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"SELECT id, thumbnail FROM products WHERE id IN ({placeholders})", ids)
        products_rows = cursor.fetchall()

        for prod in products_rows:
            thumb = prod[1] or ''
            if thumb:
                static_path = os.path.join('static', thumb.lstrip('/'))
                if os.path.exists(static_path):
                    try:
                        os.remove(static_path)
                    except Exception:
                        pass

        cursor.execute(f"DELETE FROM products WHERE id IN ({placeholders})", ids)
        conn.commit()
        return jsonify({'success': True, 'deleted': len(ids)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/products/export')
@admin_required
def export_products():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.id, p.name, p.name_am, p.name_ar, p.price, p.compare_price,
                   p.stock_quantity, c.name as category_name, p.is_featured, p.created_at
            FROM products p LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.is_active = 1 ORDER BY p.id DESC
        """)
        products_rows = cursor.fetchall()

        def _csv_safe(value):
            """Prefix formula-injection triggers with a single quote so
            spreadsheet apps (Excel, Sheets) never evaluate them as formulas.
            Strip leading whitespace/newlines first so they cannot be used to
            hide a trigger character from the check."""
            s = str(value) if value is not None else ''
            s_stripped = s.lstrip(' \t\r\n')
            if s_stripped and s_stripped[0] in ('=', '+', '-', '@', '\t', '\r', '\n'):
                s = "'" + s
            return s

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['ID', 'Name (EN)', 'Name (AM)', 'Name (AR)', 'Price', 'Compare Price',
                         'Stock', 'Category', 'Featured', 'Created Date'])
        for p in products_rows:
            writer.writerow([
                p['id'],
                _csv_safe(p['name']), _csv_safe(p['name_am']), _csv_safe(p['name_ar']),
                p['price'], p['compare_price'] or '', p['stock_quantity'],
                _csv_safe(p['category_name']), 'Yes' if p['is_featured'] else 'No', p['created_at']
            ])

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = (
            f'attachment; filename=products_export_{datetime_.datetime.now().strftime("%Y%m%d")}.csv'
        )
        return response
    except Exception as e:
        current_app.logger.error(f"Export products error: {e}")
        flash('Error exporting products', 'error')
        return redirect(url_for('admin.products'))


@admin_bp.route('/products/import/sample')
@admin_required
def import_products_sample():
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['name_en', 'name_am', 'name_ar', 'price', 'compare_price', 'stock',
                     'category', 'description_en', 'description_am', 'featured', 'is_new',
                     'sku', 'image_url'])
    writer.writerow(['Luxury Sofa', 'ቅርጫ ሶፋ', 'صوفا فاخرة', '15000', '20000', '10',
                     'Sofa', 'Premium quality sofa', 'ከፍተኛ ጥራት ያለው ሶፋ', 'yes', 'yes',
                     'SKU-001', 'https://images.unsplash.com/photo-1555041469-a586c61ea9bc%sw=400'])
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename=products_import_sample.csv'
    return response


@admin_bp.route('/products/import', methods=['GET', 'POST'])
@admin_required
def import_products():
    lang = get_lang()
    import secrets as _sec
    if 'import_csrf' not in session:
        session['import_csrf'] = _sec.token_hex(32)
    csrf_token = session['import_csrf']

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, name_am FROM categories WHERE is_active = 1 ORDER BY sort_order")
    categories_list = [dict(c) for c in cursor.fetchall()]

    cat_lookup = {}
    for c in categories_list:
        cat_lookup[str(c['id'])] = c['id']
        if c.get('name_am'):
            cat_lookup[c['name_am'].strip().lower()] = c['id']
        if c.get('name'):
            cat_lookup[c['name'].strip().lower()] = c['id']

    if request.method == 'GET':
        return render_template('admin/products/import.html',
                               categories=categories_list, import_csrf=csrf_token, lang=lang)

    submitted_token = request.form.get('import_csrf_token', '')
    if submitted_token != session.get('import_csrf', ''):
        flash('Invalid or expired security token. Please try again.', 'error')
        return redirect(url_for('admin.import_products'))
    session.pop('import_csrf', None)

    csv_file = request.files.get('csv_file')
    if not csv_file or csv_file.filename == '':
        flash('Please select a CSV or Excel file to upload.', 'error')
        return redirect(url_for('admin.import_products'))

    fname_lower = csv_file.filename.lower()
    is_excel = fname_lower.endswith('.xlsx') or fname_lower.endswith('.xls')
    is_csv   = fname_lower.endswith('.csv')

    if not is_csv and not is_excel:
        flash('Invalid file type. Please upload a .csv or .xlsx file.', 'error')
        return redirect(url_for('admin.import_products'))

    rows_dicts = []
    try:
        raw = csv_file.read()
        if is_excel:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                flash('The Excel file appears to be empty.', 'error')
                return redirect(url_for('admin.import_products'))
            headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
            for data_row in all_rows[1:]:
                row_dict = {headers[i]: (str(v).strip() if v is not None else '')
                            for i, v in enumerate(data_row) if i < len(headers)}
                rows_dicts.append(row_dict)
            wb.close()
        else:
            try:
                content = raw.decode('utf-8-sig')
            except UnicodeDecodeError:
                content = raw.decode('latin-1')
            reader = csv.DictReader(StringIO(content))
            if reader.fieldnames is None:
                flash('The CSV file appears to be empty.', 'error')
                return redirect(url_for('admin.import_products'))
            reader.fieldnames = [h.strip() for h in reader.fieldnames]
            rows_dicts = list(reader)
            headers = reader.fieldnames
    except Exception as ex:
        flash(f'Could not read the uploaded file: {ex}', 'error')
        return redirect(url_for('admin.import_products'))

    if not rows_dicts:
        flash('The file appears to be empty.', 'error')
        return redirect(url_for('admin.import_products'))

    headers_norm = {k.strip().lower() for k in (rows_dicts[0].keys() if rows_dicts else [])}
    REQUIRED = {'name_en', 'price', 'category'}
    missing_cols = REQUIRED - headers_norm
    if missing_cols:
        flash(f'Missing required columns: {", ".join(sorted(missing_cols))}.', 'error')
        return redirect(url_for('admin.import_products'))

    # Normalize all keys to lowercase
    rows_dicts = [{k.strip().lower(): v for k, v in r.items()} for r in rows_dicts]
    upload_dir = 'static/uploads/products'
    os.makedirs(upload_dir, exist_ok=True)

    validated_rows = []
    row_errors = []

    def _bool_field(val):
        return 1 if val.strip().lower() in ('yes', '1', 'true') else 0

    for row_num, row in enumerate(rows_dicts, start=2):
        row = {k: (v or '').strip() for k, v in row.items() if k}

        name_en = row.get('name_en', '')
        if not name_en:
            row_errors.append({'row': row_num, 'field': 'name_en', 'message': '"name_en" is required.'})
            continue

        price_raw = row.get('price', '')
        try:
            price = float(price_raw)
            if price <= 0:
                raise ValueError
        except (ValueError, TypeError):
            row_errors.append({'row': row_num, 'field': 'price',
                               'message': f'Must be a positive number (got "{price_raw}").'})
            continue

        compare_price_raw = row.get('compare_price', '')
        compare_price = None
        if compare_price_raw:
            try:
                compare_price = float(compare_price_raw)
            except (ValueError, TypeError):
                row_errors.append({'row': row_num, 'field': 'compare_price',
                                   'message': f'Must be a number (got "{compare_price_raw}").'})
                continue

        stock_raw = row.get('stock', '').strip()
        try:
            stock = int(float(stock_raw)) if stock_raw else 99
        except (ValueError, TypeError):
            stock = 99

        category_raw = row.get('category', '').strip()
        category_id = cat_lookup.get(category_raw.lower()) or cat_lookup.get(category_raw)
        if category_id is None:
            row_errors.append({'row': row_num, 'field': 'category',
                               'message': f'Unknown category "{category_raw}".'})
            continue

        validated_rows.append({
            'name': name_en,
            'name_en': name_en,
            'name_am': row.get('name_am', ''),
            'name_ar': row.get('name_ar', ''),
            'price': price,
            'compare_price': compare_price,
            'stock_quantity': stock,
            'category_id': category_id,
            'description': row.get('description_en', ''),
            'description_en': row.get('description_en', ''),
            'description_am': row.get('description_am', ''),
            'sku': row.get('sku', ''),
            'is_featured': _bool_field(row.get('featured', '0')),
            'is_new': _bool_field(row.get('is_new', '0')),
            'image_url': row.get('image_url', '').strip(),
        })

    if row_errors and not validated_rows:
        flash(f'No valid rows found. {len(row_errors)} error(s) detected.', 'error')
        return render_template('admin/products/import.html',
                               categories=categories_list, import_csrf=csrf_token,
                               row_errors=row_errors, lang=lang)

    imported = 0
    for vrow in validated_rows:
        thumbnail = ''
        if vrow['image_url']:
            try:
                import requests as _req
                import socket as _socket
                from urllib.parse import urlparse as _urlparse
                import ipaddress as _ipaddress

                def _assert_safe_host(hostname):
                    """Resolve hostname and reject any non-global IP (SSRF guard)."""
                    try:
                        infos = _socket.getaddrinfo(hostname, None)
                    except _socket.gaierror:
                        raise ValueError(f"Cannot resolve hostname: {hostname}")
                    for info in infos:
                        addr = info[4][0]
                        try:
                            ip = _ipaddress.ip_address(addr)
                            if not ip.is_global:
                                raise ValueError(
                                    f"Image URL resolves to a private/internal address ({addr})"
                                )
                        except ValueError as _ve:
                            if 'private' in str(_ve) or 'internal' in str(_ve):
                                raise

                _parsed = _urlparse(vrow['image_url'])
                # Block non-HTTP(S) schemes (SSRF guard)
                if _parsed.scheme not in ('http', 'https'):
                    raise ValueError("Only http/https image URLs are allowed")
                _assert_safe_host(_parsed.hostname or '')
                # Disable redirects so a redirect to an internal URL cannot bypass the check
                resp = _req.get(vrow['image_url'], timeout=10, stream=True,
                                allow_redirects=False)
                if resp.status_code == 200:
                    ct = resp.headers.get('Content-Type', '').split(';')[0].strip()
                    ext_map = {'image/jpeg': 'jpg', 'image/png': 'png',
                               'image/webp': 'webp', 'image/gif': 'gif'}
                    ext = ext_map.get(ct, 'jpg')
                    fname = f"import_{uuid.uuid4().hex}.{ext}"
                    with open(os.path.join(upload_dir, fname), 'wb') as f:
                        for chunk in resp.iter_content(chunk_size=65536):
                            f.write(chunk)
                    thumbnail = f'uploads/products/{fname}'
            except Exception:
                pass

        try:
            cursor.execute("""
                INSERT INTO products (
                    name, name_en, name_am, name_ar,
                    description, description_en, description_am,
                    price, compare_price, stock_quantity, category_id,
                    sku, is_featured, is_new, thumbnail, is_active
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
            """, (vrow['name'], vrow['name_en'], vrow['name_am'], vrow['name_ar'],
                  vrow['description'], vrow['description_en'], vrow['description_am'],
                  vrow['price'], vrow['compare_price'], vrow['stock_quantity'], vrow['category_id'],
                  vrow['sku'], vrow['is_featured'], vrow['is_new'], thumbnail))
            imported += 1
        except Exception as e:
            current_app.logger.error(f"Import row error: {e}")

    conn.commit()
    flash(f'Successfully imported {imported} product(s). {len(row_errors)} row(s) skipped.', 'success')
    return redirect(url_for('admin.products'))


# ==================== ADVERTISEMENT MANAGEMENT ====================

@admin_bp.route('/ads')
@admin_required
def ads():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM advertisements ORDER BY sort_order ASC, id DESC")
        ads_list = cursor.fetchall()
        return render_template('admin/ads/index.html',
                               ads=[dict(a) for a in ads_list] if ads_list else [], lang=lang)
    except Exception as e:
        current_app.logger.error(f"Admin ads error: {e}")
        flash('Error loading ads.', 'error')
        return redirect(url_for('admin.dashboard'))


@admin_bp.route('/ads/create', methods=['GET', 'POST'])
@admin_required
def ad_create():
    if request.method == 'POST':
        title = request.form.get('title', '')
        title_am = request.form.get('title_am', '')
        title_ar = request.form.get('title_ar', '')
        description = request.form.get('ad_text', '') or request.form.get('description', '')
        description_am = request.form.get('description_am', '')
        description_ar = request.form.get('description_ar', '')
        link = request.form.get('link', '')
        try:
            sort_order = int(request.form.get('sort_order', 0) or 0)
        except (ValueError, TypeError):
            flash('Sort order must be a number.', 'error')
            return redirect(url_for('admin.ad_create'))

        _ALLOWED_IMG_EXT = {'jpg', 'jpeg', 'png', 'webp', 'gif'}
        _ALLOWED_VID_EXT = {'mp4', 'webm', 'mov', 'avi', 'ogg'}
        _ALLOWED_AD_EXT  = _ALLOWED_IMG_EXT | _ALLOWED_VID_EXT
        image_filename = ''
        image = request.files.get('image')
        if image and image.filename:
            from werkzeug.utils import secure_filename
            import subprocess as _sp
            filename = secure_filename(image.filename)
            ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            if ext not in _ALLOWED_AD_EXT:
                flash(f"File type '.{ext}' is not allowed. Use jpg/png/webp/gif or mp4/webm/mov.", 'error')
                return redirect(url_for('admin.ad_create'))
            upload_dir = 'static/uploads/ads'
            os.makedirs(upload_dir, exist_ok=True)
            unique_filename = f"ad_{uuid.uuid4().hex[:8]}.{ext}"
            saved_path = os.path.join(upload_dir, unique_filename)
            image.save(saved_path)
            if ext in _ALLOWED_VID_EXT:
                base_id = uuid.uuid4().hex[:8]
                converted_name = f"ad_{base_id}.mp4"
                converted_path = os.path.join(upload_dir, converted_name)
                poster_name    = f"ad_{base_id}_poster.jpg"
                poster_path    = os.path.join(upload_dir, poster_name)
                try:
                    r = _sp.run(
                        ['ffmpeg', '-i', saved_path,
                         '-c:v', 'libx264', '-preset', 'fast', '-crf', '23',
                         '-c:a', 'aac', '-b:a', '128k',
                         '-movflags', '+faststart', '-pix_fmt', 'yuv420p',
                         converted_path, '-y'],
                        capture_output=True, timeout=120
                    )
                    if r.returncode == 0:
                        os.remove(saved_path)
                        image_filename = f'uploads/ads/{converted_name}'
                        _sp.run(['ffmpeg', '-i', converted_path,
                                 '-ss', '00:00:01', '-vframes', '1',
                                 '-update', '1', poster_path, '-y'],
                                capture_output=True, timeout=30)
                    else:
                        image_filename = f'uploads/ads/{unique_filename}'
                except Exception:
                    image_filename = f'uploads/ads/{unique_filename}'
            else:
                image_filename = f'uploads/ads/{unique_filename}'

        media_url = request.form.get('media_url', '').strip()

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO advertisements (
                title, title_am, title_ar, description, description_am, description_ar,
                image, media_url, link, sort_order, is_active
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
        """, (title, title_am, title_ar, description, description_am, description_ar,
              image_filename, media_url, link, sort_order))
        conn.commit()
        flash('Advertisement created successfully!', 'success')
        return redirect(url_for('admin.ads'))

    return render_template('admin/ads/create.html')


@admin_bp.route('/ads/edit/<int:aid>', methods=['GET', 'POST'])
@admin_required
def ad_edit(aid):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM advertisements WHERE id = %s", (aid,))
    ad = cursor.fetchone()
    if not ad:
        flash('Advertisement not found!', 'danger')
        return redirect(url_for('admin.ads'))

    if request.method == 'POST':
        title = request.form.get('title', '')
        title_am = request.form.get('title_am', '')
        title_ar = request.form.get('title_ar', '')
        description = request.form.get('ad_text', '') or request.form.get('description', '')
        description_am = request.form.get('description_am', '')
        description_ar = request.form.get('description_ar', '')
        link = request.form.get('link', '')
        is_active = 1 if request.form.get('is_active') else 0
        try:
            sort_order = int(float(request.form.get('sort_order', 0) or 0))
        except (ValueError, TypeError):
            flash('Sort order must be a number.', 'error')
            return redirect(url_for('admin.ad_edit', aid=aid))

        _ALLOWED_IMG_EXT = {'jpg', 'jpeg', 'png', 'webp', 'gif'}
        _ALLOWED_VID_EXT = {'mp4', 'webm', 'mov', 'avi', 'ogg'}
        _ALLOWED_AD_EXT  = _ALLOWED_IMG_EXT | _ALLOWED_VID_EXT
        image_filename = ad['image']
        image = request.files.get('image')
        if image and image.filename:
            from werkzeug.utils import secure_filename
            import subprocess as _sp
            filename = secure_filename(image.filename)
            ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            if ext not in _ALLOWED_AD_EXT:
                flash(f"File type '.{ext}' is not allowed. Use jpg/png/webp/gif or mp4/webm/mov.", 'error')
                return redirect(url_for('admin.ad_edit', aid=aid))
            upload_dir = 'static/uploads/ads'
            os.makedirs(upload_dir, exist_ok=True)
            unique_filename = f"ad_{uuid.uuid4().hex[:8]}.{ext}"
            saved_path = os.path.join(upload_dir, unique_filename)
            image.save(saved_path)
            if ext in _ALLOWED_VID_EXT:
                base_id = uuid.uuid4().hex[:8]
                converted_name = f"ad_{base_id}.mp4"
                converted_path = os.path.join(upload_dir, converted_name)
                poster_name    = f"ad_{base_id}_poster.jpg"
                poster_path    = os.path.join(upload_dir, poster_name)
                try:
                    r = _sp.run(
                        ['ffmpeg', '-i', saved_path,
                         '-c:v', 'libx264', '-preset', 'fast', '-crf', '23',
                         '-c:a', 'aac', '-b:a', '128k',
                         '-movflags', '+faststart', '-pix_fmt', 'yuv420p',
                         converted_path, '-y'],
                        capture_output=True, timeout=120
                    )
                    if r.returncode == 0:
                        os.remove(saved_path)
                        image_filename = f'uploads/ads/{converted_name}'
                        _sp.run(['ffmpeg', '-i', converted_path,
                                 '-ss', '00:00:01', '-vframes', '1',
                                 '-update', '1', poster_path, '-y'],
                                capture_output=True, timeout=30)
                    else:
                        image_filename = f'uploads/ads/{unique_filename}'
                except Exception:
                    image_filename = f'uploads/ads/{unique_filename}'
            else:
                image_filename = f'uploads/ads/{unique_filename}'

        media_url = request.form.get('media_url', ad.get('media_url', '') or '').strip()

        try:
            cursor.execute("""
                UPDATE advertisements SET
                    title=%s, title_am=%s, title_ar=%s, description=%s, description_am=%s, description_ar=%s,
                    image=%s, media_url=%s, link=%s, sort_order=%s, is_active=%s
                WHERE id=%s
            """, (title, title_am, title_ar, description, description_am, description_ar,
                  image_filename, media_url, link, sort_order, is_active, aid))
            conn.commit()
        except Exception as _db_err:
            conn.rollback()
            current_app.logger.error(f"Ad edit DB error for ad {aid}: {_db_err}")
            flash(f'Database error saving advertisement: {_db_err}', 'error')
            return redirect(url_for('admin.ad_edit', aid=aid))
        flash('Advertisement updated successfully!', 'success')
        return redirect(url_for('admin.ads'))

    return render_template('admin/ads/edit.html', ad=dict(ad))


@admin_bp.route('/ads/toggle/<int:aid>', methods=['POST'])
@admin_required
def ad_toggle(aid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE advertisements SET is_active = NOT is_active WHERE id = %s RETURNING is_active",
            (aid,)
        )
        row = cursor.fetchone()
        conn.commit()
        new_status = bool(row[0]) if row else None
        return jsonify({'success': True, 'is_active': new_status})
    except Exception as e:
        current_app.logger.error(f"Ad toggle error: {e}")
        return jsonify({'success': False, 'error': 'Could not toggle advertisement status'}), 500


@admin_bp.route('/ads/delete/<int:aid>', methods=['POST', 'DELETE'])
@admin_required
def ad_delete(aid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM advertisements WHERE id = %s", (aid,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Advertisement deleted successfully'})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        current_app.logger.error("Ad delete error: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Could not delete advertisement'}), 500


@admin_bp.route('/ads/clear-media/<int:aid>', methods=['POST'])
@admin_required
def ad_clear_media(aid):
    """Remove only the image/video from an ad without deleting the ad itself."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("UPDATE advertisements SET image = '', media_url = '' WHERE id = %s", (aid,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Media removed from advertisement'})
    except Exception as e:
        conn.rollback()
        current_app.logger.error(f"Ad clear media error for ad {aid}: {e}")
        return jsonify({'success': False, 'error': 'Could not remove media'}), 500


@admin_bp.route('/ads/reorder', methods=['POST'])
@admin_required
def ad_reorder():
    try:
        data = request.get_json()
        src_id = data.get('src_id')
        dest_id = data.get('dest_id')
        if not src_id or not dest_id:
            return jsonify({'success': False, 'error': 'Invalid request'}), 400

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT sort_order FROM advertisements WHERE id = %s", (src_id,))
        src_order = cursor.fetchone()
        cursor.execute("SELECT sort_order FROM advertisements WHERE id = %s", (dest_id,))
        dest_order = cursor.fetchone()
        if src_order and dest_order:
            cursor.execute("UPDATE advertisements SET sort_order = %s WHERE id = %s", (dest_order[0], src_id))
            cursor.execute("UPDATE advertisements SET sort_order = %s WHERE id = %s", (src_order[0], dest_id))
            conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== ORDER MANAGEMENT ====================

@admin_bp.route('/orders')
@admin_required
def orders():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()

        status = request.args.get('status', 'all')
        search = request.args.get('search', '').strip()

        query = """
            SELECT o.*, u.full_name as customer_name
            FROM orders o LEFT JOIN users u ON o.user_id = u.id WHERE 1=1
        """
        params = []
        if status != 'all':
            query += " AND o.status = %s"
            params.append(status)
        if search:
            query += " AND (o.order_number ILIKE %s OR u.full_name ILIKE %s OR o.shipping_phone ILIKE %s)"
            s = f'%{search}%'
            params.extend([s, s, s])
        query += " ORDER BY o.id DESC"

        cursor.execute(query, params)
        orders_rows = cursor.fetchall()
        orders_list = [dict(o) for o in orders_rows] if orders_rows else []

        cursor.execute("SELECT status, COUNT(*) as count FROM orders GROUP BY status")
        status_counts = {row['status']: row['count'] for row in cursor.fetchall()}

        return render_template('admin/orders/index.html',
                               orders=orders_list, status_counts=status_counts,
                               current_status=status, search=search, lang=lang)
    except Exception as e:
        current_app.logger.error(f"Admin orders error: {e}")
        flash('Error loading orders.', 'error')
        return redirect(url_for('admin.dashboard'))


@admin_bp.route('/orders/<int:oid>')
@admin_required
def order_detail(oid):
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT o.*, u.full_name, u.email, u.phone
            FROM orders o LEFT JOIN users u ON o.user_id = u.id
            WHERE o.id = %s
        """, (oid,))
        order = cursor.fetchone()
        if not order:
            flash('Order not found!', 'danger')
            return redirect(url_for('admin.orders'))

        cursor.execute("""
            SELECT oi.*, p.name, p.name_am, p.name_ar, p.thumbnail
            FROM order_items oi JOIN products p ON oi.product_id = p.id
            WHERE oi.order_id = %s
        """, (oid,))
        items = cursor.fetchall()

        return render_template('admin/orders/detail.html',
                               order=dict(order),
                               order_items=[dict(i) for i in items] if items else [],
                               lang=lang)
    except Exception as e:
        current_app.logger.error(f"Order detail error: {e}")
        flash('Error loading order details.', 'error')
        return redirect(url_for('admin.orders'))


@admin_bp.route('/orders/update/<int:oid>', methods=['POST'])
@admin_required
def order_update_status(oid):
    try:
        status = request.form.get('status', 'pending')
        valid_statuses = ['pending', 'confirmed', 'processing', 'shipped', 'delivered', 'cancelled']
        if status not in valid_statuses:
            return jsonify({'success': False, 'error': 'Invalid status'}), 400

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("SELECT status, user_id, total, order_number FROM orders WHERE id = %s", (oid,))
        prev = cursor.fetchone()
        prev_status   = prev[0] if prev else None
        user_id       = prev[1] if prev else None
        order_total   = float(prev[2]) if prev and prev[2] else 0
        order_number  = prev[3] if prev else str(oid)

        cursor.execute("UPDATE orders SET status = %s, updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                       (status, oid))

        # Record status change in history
        if prev_status != status:
            status_note = request.form.get('status_note', '').strip() or None
            try:
                cursor.execute(
                    "INSERT INTO order_status_history (order_id, status, note) VALUES (%s, %s, %s)",
                    (oid, status, status_note)
                )
            except Exception as _he:
                current_app.logger.warning(f"Could not record status history: {_he}")

        if status == 'delivered' and prev_status != 'delivered' and user_id:
            points_earned = max(1, int(order_total // 100))
            cursor.execute(
                "UPDATE users SET loyalty_points = COALESCE(loyalty_points, 0) + %s WHERE id = %s",
                (points_earned, user_id)
            )
            cursor.execute(
                "INSERT INTO loyalty_transactions (user_id, order_id, points, type, description) VALUES (%s, %s, %s, 'earn', %s)",
                (user_id, oid, points_earned, f'Points earned for order #{oid}')
            )

        conn.commit()

        status_icons = {
            'confirmed':  '✅', 'processing': '⚙️', 'shipped': '🚚',
            'delivered':  '🎉', 'cancelled':  '❌', 'pending': '⏳'
        }
        status_msgs = {
            'confirmed':  'Your order has been confirmed and is being prepared.',
            'processing': 'Your order is being processed by our team.',
            'shipped':    'Your order is on its way! Expect delivery soon.',
            'delivered':  f'Your order has been delivered! You earned {max(1,int(order_total//100)) if order_total else 1} loyalty points.',
            'cancelled':  'Your order has been cancelled. Contact us if you have questions.',
            'pending':    'Your order is pending review.',
        }
        wa_notify_url = None
        auto_sent = False

        # In-app notification for registered users only
        if user_id and prev_status != status:
            try:
                icon = status_icons.get(status, '📦')
                msg  = status_msgs.get(status, f'Order status updated to {status}.')
                notify_user(
                    user_id,
                    f'{icon} Order #{order_number} — {status.title()}',
                    msg,
                    type='order',
                    link=f'/track-order/{order_number}'
                )
            except Exception:
                pass

        # WhatsApp customer notification (auto + manual fallback)
        if prev_status != status:
            try:
                cursor.execute(
                    "SELECT shipping_phone, customer_name FROM orders WHERE id = %s", (oid,)
                )
                orow = cursor.fetchone()
                cust_phone = (orow[0] if orow else '') or ''
                cust_name  = (orow[1] if orow else '') or 'ደንበኛ'

                from services.whatsapp_service import send_customer_status_notification
                track_url = url_for('customer.track_order_public', order=order_number, _external=True)
                notif = send_customer_status_notification(
                    order_number=order_number,
                    customer_name=cust_name,
                    customer_phone=cust_phone,
                    status=status,
                    notes=f'ትዕዛዝ ይከታተሉ: {track_url}',
                )
                auto_sent     = notif.get('auto_sent', False)
                wa_notify_url = notif.get('wa_url')
            except Exception as _e:
                current_app.logger.error(f"WhatsApp notification error: {_e}")

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'status': status,
                'wa_notify_url': wa_notify_url,
                'auto_sent': auto_sent,
            })

        flash(f'Order status updated to {status}!', 'success')
        return redirect(url_for('admin.order_detail', oid=oid))
    except Exception as e:
        current_app.logger.error(f"Order update error: {e}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': str(e)}), 500
        flash('Error updating order status.', 'error')
        return redirect(url_for('admin.order_detail', oid=oid))


@admin_bp.route('/orders/delete/<int:oid>', methods=['POST'])
@admin_required
def delete_order(oid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM order_items WHERE order_id = %s", (oid,))
        cursor.execute("DELETE FROM orders WHERE id = %s", (oid,))
        conn.commit()
        flash('Order deleted successfully!', 'success')
    except Exception as e:
        current_app.logger.error(f"Delete order error: {e}")
        flash('Error deleting order.', 'error')
    return redirect(url_for('admin.orders'))


@admin_bp.route('/orders/export/<int:oid>')
@admin_required
def export_order(oid):
    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT o.*, u.full_name, u.email, u.phone
            FROM orders o LEFT JOIN users u ON o.user_id = u.id WHERE o.id = %s
        """, (oid,))
        order = cursor.fetchone()
        if not order:
            flash('Order not found!', 'danger')
            return redirect(url_for('admin.orders'))

        cursor.execute("""
            SELECT oi.*, p.name, p.name_am, p.name_ar
            FROM order_items oi JOIN products p ON oi.product_id = p.id
            WHERE oi.order_id = %s
        """, (oid,))
        items = cursor.fetchall()

        export_data = {
            'order': dict(order),
            'items': [dict(i) for i in items]
        }
        response = make_response(json.dumps(export_data, indent=2, default=str))
        response.headers['Content-Type'] = 'application/json'
        response.headers['Content-Disposition'] = (
            f'attachment; filename=order_{oid}_{datetime_.datetime.now().strftime("%Y%m%d")}.json'
        )
        return response
    except Exception as e:
        current_app.logger.error(f"Export order error: {e}")
        flash('Error exporting order.', 'error')
        return redirect(url_for('admin.orders'))


@admin_bp.route('/orders/export-excel')
@admin_required
def orders_export_excel():
    try:
        conn = get_db()
        cursor = conn.cursor()

        status = request.args.get('status', 'all')
        search = request.args.get('search', '').strip()

        query = """
            SELECT
                o.id, o.order_number, o.created_at, o.status,
                o.customer_name, o.shipping_phone, o.shipping_city, o.shipping_address,
                o.subtotal, o.discount, o.shipping_fee, o.total,
                o.payment_method, o.notes,
                o.customer_email,
                o.tracking_number, o.payment_status
            FROM orders o
            WHERE 1=1
        """
        params = []

        if status and status != 'all':
            query += " AND o.status = %s"
            params.append(status)

        if search:
            query += """ AND (
                o.order_number ILIKE %s OR
                o.shipping_name ILIKE %s OR
                o.shipping_phone ILIKE %s
            )"""
            like = f"%{search}%"
            params.extend([like, like, like])

        query += " ORDER BY o.created_at DESC"
        cursor.execute(query, params)
        orders = cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"

        header_fill = PatternFill(start_color="1D6F42", end_color="1D6F42", fill_type="solid")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alt_fill = PatternFill(start_color="F0F7F2", end_color="F0F7F2", fill_type="solid")
        border_side = Side(style="thin", color="CCCCCC")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        headers = [
            "ID", "Order Number", "Date", "Status",
            "Customer Name", "Phone", "City", "Address",
            "Subtotal (ETB)", "Discount (ETB)", "Shipping (ETB)", "Total (ETB)",
            "Payment Method", "Customer Email", "Notes",
            "Tracking Number", "Payment Status"
        ]

        col_widths = [6, 18, 18, 12, 22, 16, 14, 28, 14, 14, 14, 14, 16, 26, 30, 20, 16]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30

        for row_idx, order in enumerate(orders, start=2):
            is_alt = (row_idx % 2 == 0)
            row_fill = alt_fill if is_alt else None

            created_at = order['created_at']
            if created_at and hasattr(created_at, 'strftime'):
                created_at = created_at.strftime('%Y-%m-%d %H:%M')

            values = [
                order['id'],
                order['order_number'] or '',
                created_at,
                (order['status'] or 'pending').capitalize(),
                order['customer_name'] or '',
                order['shipping_phone'] or '',
                order['shipping_city'] or '',
                order['shipping_address'] or '',
                float(order['subtotal'] or 0),
                float(order['discount'] or 0),
                float(order['shipping_fee'] or 0),
                float(order['total'] or 0),
                order['payment_method'] or '',
                order['customer_email'] or '',
                order['notes'] or '',
                order['tracking_number'] or '',
                (order['payment_status'] or '').capitalize(),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = cell_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (8, 15)))
                if row_fill:
                    cell.fill = row_fill
                if col_idx in (9, 10, 11, 12):
                    cell.number_format = '#,##0.00'
                if col_idx in (8, 15):
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            ws.row_dimensions[row_idx].height = 18

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        ws2 = wb.create_sheet(title="Summary")
        ws2['A1'] = "Export Summary"
        ws2['A1'].font = Font(bold=True, size=13)
        ws2['A3'] = "Generated At:"
        ws2['B3'] = datetime_.datetime.now().strftime('%Y-%m-%d %H:%M')
        ws2['A4'] = "Filter - Status:"
        ws2['B4'] = status.capitalize()
        ws2['A5'] = "Filter - Search:"
        ws2['B5'] = search or '(none)'
        ws2['A6'] = "Total Orders:"
        ws2['B6'] = len(orders)
        ws2['A7'] = "Total Revenue (ETB):"
        ws2['B7'] = sum(float(o['total'] or 0) for o in orders)
        ws2['B7'].number_format = '#,##0.00'
        for cell in [ws2['A3'], ws2['A4'], ws2['A5'], ws2['A6'], ws2['A7']]:
            cell.font = Font(bold=True)
        ws2.column_dimensions['A'].width = 24
        ws2.column_dimensions['B'].width = 22

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        date_str = datetime_.datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"orders_{status}_{date_str}.xlsx"

        response = make_response(output.read())
        response.headers['Content-Type'] = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        current_app.logger.error(f"Export orders Excel error: {e}")
        flash('Error generating Excel export. Please try again.', 'error')
        return redirect(url_for('admin.orders'))


@admin_bp.route('/orders/invoice/<int:oid>')
@admin_required
def order_invoice(oid):
    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM orders WHERE id = %s", (oid,))
        order = cursor.fetchone()
        if not order:
            flash('Order not found!', 'danger')
            return redirect(url_for('admin.orders'))

        cursor.execute("""
            SELECT oi.*, p.name, p.name_am, p.name_ar
            FROM order_items oi JOIN products p ON oi.product_id = p.id
            WHERE oi.order_id = %s
        """, (oid,))
        items = cursor.fetchall()

        rows_html = ''.join(
            f"<tr><td>{item['name_am'] or item['name']}</td>"
            f"<td>{item['quantity']}</td>"
            f"<td>{item['price_at_time']:.2f} ETB</td>"
            f"<td>{item['price_at_time'] * item['quantity']:.2f} ETB</td></tr>"
            for item in items
        )

        html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Invoice #{order['order_number']}</title>
<style>
body{{font-family:Arial,sans-serif;margin:40px}}
.header{{text-align:center;margin-bottom:30px}}
.invoice-title{{font-size:28px;color:#1a73e8}}
table{{width:100%;border-collapse:collapse;margin:20px 0}}
th,td{{border:1px solid #ddd;padding:10px;text-align:left}}
th{{background:#1a73e8;color:white}}
.total{{font-size:18px;font-weight:bold;text-align:right}}
.footer{{margin-top:40px;text-align:center;font-size:12px;color:#666}}
</style></head>
<body>
<div class="header">
  <h1 class="invoice-title">SEMIRA FASHION</h1>
  <h2>INVOICE</h2>
</div>
<p><strong>Order #:</strong> {order['order_number']}</p>
<p><strong>Date:</strong> {order['created_at']}</p>
<p><strong>Status:</strong> {order['status']}</p>
<table>
<thead><tr><th>Product</th><th>Quantity</th><th>Price</th><th>Total</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>
<div class="total">
  <p>Subtotal: {order['subtotal']:.2f} ETB</p>
  <p>Discount: {order.get('discount', 0) or 0:.2f} ETB</p>
  <p>Shipping: {order.get('shipping_fee', 0) or 0:.2f} ETB</p>
  <p><strong>Total: {order['total']:.2f} ETB</strong></p>
</div>
<div class="footer">
  <p>Thank you for shopping with SEMIRA FASHION!</p>
  <p>ወሎ ደሴ ኩታበር, Ethiopia | +251 90 602 0606</p>
</div>
</body></html>"""
        return html
    except Exception as e:
        current_app.logger.error(f"Invoice error: {e}")
        flash('Error generating invoice.', 'error')
        return redirect(url_for('admin.order_detail', oid=oid))


# ==================== USER MANAGEMENT ====================

@admin_bp.route('/users')
@admin_required
def admin_users():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, username, email, full_name, phone, is_admin, is_active, created_at, last_login
            FROM users ORDER BY created_at DESC
        """)
        users = [dict(u) for u in cursor.fetchall()]
        return render_template('admin/users/index.html', users=users, lang=lang)
    except Exception as e:
        current_app.logger.error(f"Admin users error: {e}")
        flash('Error loading users.', 'error')
        return redirect(url_for('admin.dashboard'))


@admin_bp.route('/users/toggle/<int:uid>', methods=['POST'])
@admin_required
def toggle_user(uid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT is_active, is_admin FROM users WHERE id = %s", (uid,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        if user['is_admin']:
            return jsonify({'success': False, 'error': 'Cannot deactivate admin accounts'}), 403
        new_status = 0 if user['is_active'] else 1
        cursor.execute("UPDATE users SET is_active = %s WHERE id = %s", (new_status, uid))
        conn.commit()
        return jsonify({'success': True, 'is_active': new_status})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/users/delete/<int:uid>', methods=['POST'])
@admin_required
def delete_user(uid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT is_admin FROM users WHERE id = %s", (uid,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        if user['is_admin']:
            return jsonify({'success': False, 'error': 'Cannot delete admin accounts'}), 403
        # Remove all user-owned data before deleting the account
        cursor.execute("DELETE FROM cart_items WHERE user_id = %s", (uid,))
        cursor.execute("DELETE FROM wishlist WHERE user_id = %s", (uid,))
        cursor.execute("DELETE FROM reviews WHERE user_id = %s", (uid,))
        cursor.execute("DELETE FROM user_notifications WHERE user_id = %s", (uid,))
        cursor.execute("DELETE FROM loyalty_transactions WHERE user_id = %s", (uid,))
        # Nullify user_id on orders so order history is preserved for records
        cursor.execute("UPDATE orders SET user_id = NULL WHERE user_id = %s", (uid,))
        cursor.execute("DELETE FROM users WHERE id = %s", (uid,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/users/<int:uid>')
@admin_required
def admin_customer_detail(uid):
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id, username, email, full_name, phone, is_admin, is_active, created_at, last_login
            FROM users WHERE id = %s
        """, (uid,))
        user = cursor.fetchone()
        if not user:
            flash('User not found!', 'danger')
            return redirect(url_for('admin.admin_users'))
        user = dict(user)

        cursor.execute("""
            SELECT
                COUNT(*) AS total_orders,
                COALESCE(SUM(CASE WHEN status != 'cancelled' THEN total ELSE 0 END), 0) AS total_spent,
                COALESCE(AVG(CASE WHEN status != 'cancelled' THEN total END), 0) AS avg_order,
                COUNT(CASE WHEN status = 'pending'   THEN 1 END) AS pending_orders,
                COUNT(CASE WHEN status = 'delivered' THEN 1 END) AS delivered_orders,
                COUNT(CASE WHEN status = 'cancelled' THEN 1 END) AS cancelled_orders
            FROM orders WHERE user_id = %s
        """, (uid,))
        stats = dict(cursor.fetchone() or {})

        cursor.execute("""
            SELECT id, order_number, total, status, created_at, customer_name, shipping_phone
            FROM orders WHERE user_id = %s
            ORDER BY created_at DESC LIMIT 30
        """, (uid,))
        orders = [dict(o) for o in cursor.fetchall()]

        cursor.execute("""
            SELECT p.name, p.name_am, p.thumbnail,
                   COUNT(oi.id)       AS times_ordered,
                   SUM(oi.quantity)   AS total_qty,
                   SUM(oi.subtotal)   AS total_value
            FROM order_items oi
            JOIN orders  o ON oi.order_id  = o.id
            JOIN products p ON oi.product_id = p.id
            WHERE o.user_id = %s AND o.status != 'cancelled'
            GROUP BY p.id, p.name, p.name_am, p.thumbnail
            ORDER BY total_qty DESC LIMIT 5
        """, (uid,))
        top_products = [dict(p) for p in cursor.fetchall()]

        cursor.execute("""
            SELECT ci.quantity, ci.added_at, p.name, p.name_am, p.price, p.thumbnail
            FROM cart_items ci JOIN products p ON ci.product_id = p.id
            WHERE ci.user_id = %s ORDER BY ci.added_at DESC
        """, (uid,))
        cart_items = [dict(c) for c in cursor.fetchall()]

        return render_template('admin/users/detail.html',
                               user=user, stats=stats, orders=orders,
                               top_products=top_products, cart_items=cart_items,
                               lang=lang)
    except Exception as e:
        current_app.logger.error(f"Customer detail error: {e}")
        flash('Error loading customer details.', 'error')
        return redirect(url_for('admin.admin_users'))


# ==================== INBOX ====================

@admin_bp.route('/inbox')
@admin_required
def admin_inbox():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()

        filter_status = request.args.get('filter', 'all')
        search = request.args.get('search', '').strip()

        query = "SELECT * FROM contact_messages WHERE 1=1"
        params = []
        if filter_status == 'unread':
            query += " AND is_read = 0"
        elif filter_status == 'read':
            query += " AND is_read = 1"
        if search:
            query += " AND (name ILIKE %s OR email ILIKE %s OR phone ILIKE %s OR message ILIKE %s)"
            s = f'%{search}%'
            params.extend([s, s, s, s])
        query += " ORDER BY created_at DESC"
        cursor.execute(query, params)
        messages = [dict(r) for r in cursor.fetchall()]

        cursor.execute("SELECT COUNT(*) FROM contact_messages WHERE is_read = 0")
        unread_count = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COUNT(*) FROM contact_messages")
        total_count = cursor.fetchone()[0] or 0

        return render_template('admin/inbox/index.html',
                               messages=messages, filter_status=filter_status,
                               search=search, unread_count=unread_count,
                               total_count=total_count, lang=lang)
    except Exception as e:
        current_app.logger.error(f"Admin inbox error: {e}")
        flash('Error loading inbox.', 'error')
        return redirect(url_for('admin.dashboard'))


@admin_bp.route('/inbox/<int:mid>/mark-read', methods=['POST'])
@admin_required
def inbox_mark_read(mid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT is_read FROM contact_messages WHERE id = %s", (mid,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Message not found'}), 404
        new_status = 0 if row['is_read'] else 1
        cursor.execute("UPDATE contact_messages SET is_read = %s WHERE id = %s", (new_status, mid))
        conn.commit()
        return jsonify({'success': True, 'is_read': new_status})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/inbox/mark-all-read', methods=['POST'])
@admin_required
def inbox_mark_all_read():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("UPDATE contact_messages SET is_read = 1 WHERE is_read = 0")
        conn.commit()
        count = cursor.rowcount
        return jsonify({'success': True, 'updated': count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/inbox/<int:mid>/delete', methods=['POST'])
@admin_required
def inbox_delete(mid):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM contact_messages WHERE id = %s", (mid,))
        conn.commit()
        if cursor.rowcount == 0:
            return jsonify({'success': False, 'error': 'Message not found'}), 404
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/inbox/<int:mid>/note', methods=['POST'])
@admin_required
def inbox_save_note(mid):
    try:
        data = request.get_json(silent=True) or {}
        note = data.get('note', '').strip()
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM contact_messages WHERE id = %s", (mid,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Message not found'}), 404
        cursor.execute("UPDATE contact_messages SET admin_notes = %s WHERE id = %s",
                       (note if note else None, mid))
        conn.commit()
        return jsonify({'success': True, 'note': note})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== REPORTS ====================

@admin_bp.route('/reports')
@admin_required
def reports():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM products WHERE is_active = 1")
        total_products = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COUNT(*) FROM orders")
        total_orders = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COUNT(*) FROM users WHERE is_admin = 0")
        total_customers = cursor.fetchone()[0] or 0
        cursor.execute("SELECT SUM(total) FROM orders WHERE status = 'delivered'")
        total_revenue = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COUNT(*) FROM orders WHERE status = 'pending'")
        pending_orders = cursor.fetchone()[0] or 0
        cursor.execute("SELECT COUNT(*) FROM orders WHERE status = 'delivered'")
        completed_orders = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT c.id, c.name_am, c.name, COUNT(p.id) as product_count
            FROM categories c
            LEFT JOIN products p ON p.category_id = c.id AND p.is_active = 1
            WHERE c.is_active = 1 GROUP BY c.id, c.name_am, c.name
            ORDER BY product_count DESC
        """)
        categories = [dict(c) for c in cursor.fetchall()]

        cursor.execute("""
            SELECT TO_CHAR(created_at, 'YYYY-MM') as month,
                   COUNT(*) as order_count, SUM(total) as revenue
            FROM orders WHERE status != 'cancelled'
            AND created_at >= NOW() - INTERVAL '12 months'
            GROUP BY TO_CHAR(created_at, 'YYYY-MM')
            ORDER BY month DESC
        """)
        monthly_sales = [dict(s) for s in cursor.fetchall()]

        cursor.execute("""
            SELECT p.id, p.name_am, p.name, SUM(oi.quantity) as total_sold,
                   SUM(oi.quantity * oi.price_at_time) as revenue
            FROM order_items oi JOIN products p ON oi.product_id = p.id
            GROUP BY p.id, p.name_am, p.name
            ORDER BY total_sold DESC LIMIT 10
        """)
        top_products = [dict(p) for p in cursor.fetchall()]

        cursor.execute("""
            SELECT p.id, p.name_am, p.name, p.stock_quantity, p.low_stock_threshold
            FROM products p
            WHERE p.stock_quantity <= p.low_stock_threshold AND p.stock_quantity > 0
            ORDER BY p.stock_quantity ASC LIMIT 10
        """)
        low_stock = [dict(p) for p in cursor.fetchall()]

        stats = {
            'total_products': total_products, 'total_orders': total_orders,
            'total_customers': total_customers, 'total_revenue': total_revenue,
            'pending_orders': pending_orders, 'completed_orders': completed_orders
        }

        return render_template('admin/reports/index.html',
                               stats=stats, categories=categories,
                               monthly_sales=monthly_sales, top_products=top_products,
                               low_stock=low_stock,
                               total_products=total_products, total_orders=total_orders,
                               total_revenue=total_revenue, total_customers=total_customers,
                               pending_orders=pending_orders, completed_orders=completed_orders,
                               lang=lang)
    except Exception as e:
        current_app.logger.error(f"Reports error: {e}")
        current_app.logger.error("Traceback:", exc_info=True)
        flash('Error loading reports.', 'error')
        return redirect(url_for('admin.dashboard'))


@admin_bp.route('/reports/sales')
@admin_required
def reports_sales():
    lang = get_lang()
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    try:
        conn = get_db()
        cursor = conn.cursor()

        query = """
            SELECT o.*, u.full_name
            FROM orders o LEFT JOIN users u ON o.user_id = u.id
            WHERE o.status != 'cancelled'
        """
        params = []
        if start_date:
            query += " AND DATE(o.created_at) >= %s"
            params.append(start_date)
        if end_date:
            query += " AND DATE(o.created_at) <= %s"
            params.append(end_date)
        query += " ORDER BY o.created_at DESC"
        cursor.execute(query, params)
        orders_rows = cursor.fetchall()

        orders_list = []
        total_revenue = 0
        for o in orders_rows:
            od = dict(o)
            od['customer_name'] = od.pop('full_name', None) or 'Guest'
            od['date'] = (od.get('created_at') or '')[:10]
            total_revenue += od.get('total', 0) or 0
            orders_list.append(od)

        total_orders = len(orders_list)
        avg_order_value = total_revenue / total_orders if total_orders > 0 else 0

        cursor.execute("""
            SELECT DATE(created_at) as date, COUNT(*) as order_count, SUM(total) as revenue
            FROM orders WHERE status != 'cancelled'
            GROUP BY DATE(created_at) ORDER BY date DESC LIMIT 30
        """)
        daily_sales = [dict(s) for s in cursor.fetchall()]
        chart_labels = [s.get('date', '') for s in daily_sales]
        chart_values = [s.get('revenue') or 0 for s in daily_sales]

        cursor.execute("""
            SELECT p.id, p.name_am as name, c.name_am as category,
                   COALESCE(SUM(oi.quantity), 0) as units_sold,
                   COALESCE(SUM(oi.quantity * oi.price_at_time), 0) as revenue
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN order_items oi ON p.id = oi.product_id
            GROUP BY p.id, p.name_am, c.name_am
            ORDER BY units_sold DESC LIMIT 10
        """)
        top_products = [dict(p) for p in cursor.fetchall()]

        return render_template('admin/reports/sales.html',
                               orders=orders_list, total_revenue=total_revenue,
                               total_orders=total_orders,
                               avg_order_value=round(avg_order_value, 2),
                               daily_sales=daily_sales, top_products=top_products,
                               chart_labels=chart_labels, chart_values=chart_values,
                               start_date=start_date, end_date=end_date, lang=lang)
    except Exception as e:
        current_app.logger.error(f"Reports sales error: {e}")
        flash('Error loading sales report.', 'error')
        return redirect(url_for('admin.reports'))


@admin_bp.route('/reports/products')
@admin_required
def reports_products():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT p.id, p.name_am, p.name, p.name_ar, p.price, p.compare_price,
                   p.stock_quantity, p.low_stock_threshold, p.is_featured,
                   p.thumbnail, p.created_at, c.name as category_name,
                   COALESCE(SUM(oi.quantity), 0) as total_sold,
                   COALESCE(SUM(oi.quantity * oi.price_at_time), 0) as total_revenue
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN order_items oi ON p.id = oi.product_id
            WHERE p.is_active = 1
            GROUP BY p.id, p.name_am, p.name, p.name_ar, p.price, p.compare_price,
                     p.stock_quantity, p.low_stock_threshold, p.is_featured,
                     p.thumbnail, p.created_at, c.name
            ORDER BY total_sold DESC, p.id DESC
        """)
        products_rows = cursor.fetchall()

        cursor.execute("""
            SELECT COUNT(*) as total_products,
                   SUM(stock_quantity) as total_stock,
                   SUM(CASE WHEN stock_quantity = 0 THEN 1 ELSE 0 END) as out_of_stock,
                   SUM(CASE WHEN stock_quantity <= low_stock_threshold AND stock_quantity > 0 THEN 1 ELSE 0 END) as low_stock,
                   AVG(price) as avg_price,
                   SUM(price * stock_quantity) as total_value
            FROM products WHERE is_active = 1
        """)
        stats_row = cursor.fetchone()
        stats_dict = dict(stats_row) if stats_row else {}

        cursor.execute("""
            SELECT c.id, c.name_am, c.name, COUNT(p.id) as product_count
            FROM categories c
            LEFT JOIN products p ON p.category_id = c.id AND p.is_active = 1
            WHERE c.is_active = 1 GROUP BY c.id, c.name_am, c.name
            ORDER BY c.name_am ASC
        """)
        categories = [dict(c) for c in cursor.fetchall()]

        return render_template('admin/reports/products.html',
                               products=[dict(p) for p in products_rows] if products_rows else [],
                               stats=stats_dict, categories=categories,
                               total_products=stats_dict.get('total_products', 0),
                               total_value=stats_dict.get('total_value', 0),
                               lang=lang)
    except Exception as e:
        current_app.logger.error(f"Reports products error: {e}")
        flash('Error loading products report.', 'error')
        return redirect(url_for('admin.reports'))


# ==================== SETTINGS ====================

@admin_bp.route('/settings', methods=['GET', 'POST'])
@admin_required
def settings():
    lang = get_lang()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT key, value FROM settings")
    settings_rows = cursor.fetchall()
    settings_data = {row[0]: row[1] for row in settings_rows}

    if request.method == 'POST':
        try:
            settings_to_save = [
                ('site_name', request.form.get('site_name', 'SEMIRA FASHION')),
                ('site_description', request.form.get('site_description', '')),
                ('admin_email', request.form.get('admin_email', 'admin@semirafashion.com')),
                ('whatsapp_number', request.form.get('whatsapp_number', '251987957957')),
                ('phone_number', request.form.get('phone_number', '+251987957957')),
                ('store_address', request.form.get('store_address', 'ወሎ ደሴ ኩታበር, Ethiopia')),
                ('free_shipping_threshold', request.form.get('free_shipping_threshold', '5000')),
                ('shipping_cost', request.form.get('shipping_cost', '200')),
                ('currency', request.form.get('currency', 'ETB')),
                ('default_language', request.form.get('default_language', 'am')),
                ('meta_keywords', request.form.get('meta_keywords', '')),
                ('google_analytics', request.form.get('google_analytics', '')),
            ]
            for key, value in settings_to_save:
                cursor.execute("""
                    INSERT INTO settings (key, value) VALUES (%s, %s)
                    ON CONFLICT(key) DO UPDATE SET value = %s
                """, (key, value, value))
            conn.commit()
            flash('Settings saved successfully!', 'success')
        except Exception as e:
            current_app.logger.error(f"Settings save error: {e}")
            flash('Error saving settings.', 'error')
        return redirect(url_for('admin.settings'))

    groq_configured = bool(
        os.environ.get('GROQ_API_KEY') or settings_data.get('groq_api_key')
    )
    return render_template('admin/settings.html', settings=settings_data,
                           lang=lang, groq_configured=groq_configured)


@admin_bp.route('/settings/change-password', methods=['POST'])
@admin_required
def change_password():
    current_password = request.form.get('current_password', '').strip()
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if not current_password or not new_password or not confirm_password:
        flash('All password fields are required.', 'error')
        return redirect(url_for('admin.settings') + '#change-password')

    if new_password != confirm_password:
        flash('New password and confirmation do not match.', 'error')
        return redirect(url_for('admin.settings') + '#change-password')

    if len(new_password) < 8:
        flash('New password must be at least 8 characters long.', 'error')
        return redirect(url_for('admin.settings') + '#change-password')

    try:
        current_admin_password = os.environ.get('ADMIN_PASSWORD', '1234')

        # Check settings table for a stored hash override first
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT value FROM settings WHERE key = 'admin_password_hash'")
        row = cursor.fetchone()
        stored_hash = row[0] if row else None

        # Verify current password
        if stored_hash:
            password_ok = check_password_hash(stored_hash, current_password)
        else:
            password_ok = (current_password == current_admin_password)

        if not password_ok:
            flash('Current password is incorrect.', 'error')
            return redirect(url_for('admin.settings') + '#change-password')

        new_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
        # Store new hash in settings table for persistence across restarts
        cursor.execute("""
            INSERT INTO settings (key, value) VALUES ('admin_password_hash', %s)
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
        """, (new_hash,))
        conn.commit()
        # Also update in-memory env var for this session
        os.environ['ADMIN_PASSWORD'] = new_password
        flash('Password changed successfully! Update ADMIN_PASSWORD in Replit Secrets to make it permanent.', 'success')
    except Exception as e:
        current_app.logger.error(f"Admin change password error: {e}")
        flash('Error changing password. Please try again.', 'error')

    return redirect(url_for('admin.settings') + '#change-password')


@admin_bp.route('/ai-logs')
@admin_required
def ai_logs():
    """AI conversation log — paginated list with stats and filters."""
    lang = get_lang()
    conn = get_db()
    cursor = conn.cursor()

    try:
        page = max(1, int(request.args.get('page', 1)))
    except (ValueError, TypeError):
        page = 1
    per_page  = 20
    offset    = (page - 1) * per_page
    source_f  = request.args.get('source', '')   # groq | fallback | error | ''
    lang_f    = request.args.get('lang', '')
    q         = request.args.get('q', '').strip()
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    # Build WHERE clause
    wheres = []
    params = []
    if source_f:
        wheres.append("source = %s"); params.append(source_f)
    if lang_f:
        wheres.append("lang = %s"); params.append(lang_f)
    if q:
        wheres.append("(user_message ILIKE %s OR ai_reply ILIKE %s OR user_name ILIKE %s)")
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if date_from:
        wheres.append("created_at::date >= %s"); params.append(date_from)
    if date_to:
        wheres.append("created_at::date <= %s"); params.append(date_to)

    where_sql = ('WHERE ' + ' AND '.join(wheres)) if wheres else ''

    # Total count
    cursor.execute(f"SELECT COUNT(*) FROM ai_conversations {where_sql}", params)
    total = cursor.fetchone()[0]
    total_pages = max(1, -(-total // per_page))

    # Rows
    cursor.execute(f"""
        SELECT id, user_id, user_name, user_message, ai_reply, source, lang, ip_address, created_at
        FROM ai_conversations {where_sql}
        ORDER BY created_at DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    rows = cursor.fetchall()

    # Summary stats (always over full table)
    cursor.execute("""
        SELECT
            COUNT(*) AS total,
            COUNT(*) FILTER (WHERE source = 'groq')     AS groq_count,
            COUNT(*) FILTER (WHERE source = 'fallback') AS fallback_count,
            COUNT(*) FILTER (WHERE source = 'error')    AS error_count,
            COUNT(DISTINCT ip_address)                   AS unique_ips,
            COUNT(*) FILTER (WHERE created_at >= NOW() - INTERVAL '24 hours') AS today
        FROM ai_conversations
    """)
    stats = cursor.fetchone()

    # Top 5 questions
    cursor.execute("""
        SELECT user_message, COUNT(*) AS n
        FROM ai_conversations
        GROUP BY user_message
        ORDER BY n DESC
        LIMIT 5
    """)
    top_questions = cursor.fetchall()

    return render_template(
        'admin/ai_logs.html',
        lang=lang,
        rows=rows,
        stats=stats,
        top_questions=top_questions,
        page=page,
        total_pages=total_pages,
        total=total,
        source_f=source_f,
        lang_f=lang_f,
        q=q,
        date_from=date_from,
        date_to=date_to,
    )


@admin_bp.route('/settings/smtp', methods=['POST'])
@admin_required
def save_smtp_settings():
    """Save SMTP credentials to env + settings table."""
    try:
        conn   = get_db()
        cursor = conn.cursor()

        smtp_user = request.form.get('smtp_user', '').strip()
        smtp_host = request.form.get('smtp_host', 'smtp.gmail.com').strip()
        smtp_port = request.form.get('smtp_port', '587').strip()
        smtp_pass_raw = request.form.get('smtp_pass', '').strip()

        # Only update password if user typed something new (not the placeholder)
        update_pass = smtp_pass_raw and smtp_pass_raw != '••••••••'

        pairs = [
            ('smtp_user', smtp_user),
            ('smtp_host', smtp_host),
            ('smtp_port', smtp_port),
        ]
        if update_pass:
            pairs.append(('smtp_pass', smtp_pass_raw))

        for key, val in pairs:
            cursor.execute("""
                INSERT INTO settings (key, value)
                VALUES (%s, %s)
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
            """, (key, val))
            os.environ[key.upper()] = val

        conn.commit()

        # If password saved, also load it into env
        if update_pass:
            os.environ['SMTP_PASS'] = smtp_pass_raw

        flash('SMTP settings saved successfully! "Test Email Now" ን ተጠቀሙ።', 'success')
    except Exception as e:
        current_app.logger.error(f"save_smtp_settings error: {e}")
        flash('Error saving SMTP settings.', 'error')
    return redirect(url_for('admin.settings') + '#email-digest')


@admin_bp.route('/settings/send-digest', methods=['POST'])
@admin_required
def send_test_digest():
    """Send a test digest email immediately (today's real data)."""
    try:
        from services.email_service import (
            is_email_configured, send_email,
            build_digest_html, build_digest_text
        )
        import datetime as _dt

        # Load SMTP from DB settings into env (in case just saved)
        conn   = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT key, value FROM settings WHERE key IN ('smtp_user','smtp_pass','smtp_host','smtp_port')")
        for row in cursor.fetchall():
            if row['value']:
                os.environ[row['key'].upper()] = row['value']

        if not is_email_configured():
            return jsonify({'ok': False, 'msg': 'SMTP_USER / SMTP_PASS አልተዋቀረም። Settings-ወ ላይ ያስቀምጡ።'})

        today = _dt.date.today()

        cursor.execute(
            "SELECT COUNT(*), COALESCE(SUM(total),0) FROM orders WHERE created_at::date = %s",
            (today,)
        )
        row = cursor.fetchone()
        new_orders, total_revenue = row[0] or 0, float(row[1] or 0)

        cursor.execute("SELECT COUNT(*) FROM orders WHERE status='pending'")
        pending_orders = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT name, name_am, stock_quantity, low_stock_threshold
            FROM products WHERE is_active=1 AND stock_quantity <= low_stock_threshold
            ORDER BY stock_quantity ASC LIMIT 10
        """)
        low_stock = [dict(zip(['name','name_am','stock_quantity','low_stock_threshold'], r))
                     for r in cursor.fetchall()]

        cursor.execute(
            "SELECT COUNT(*) FROM ai_conversations WHERE created_at::date = %s", (today,)
        )
        ai_conversations = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT p.name, p.name_am, COUNT(oi.id) AS order_count
            FROM order_items oi
            JOIN orders o ON o.id = oi.order_id
            JOIN products p ON p.id = oi.product_id
            WHERE o.created_at::date = %s
            GROUP BY p.id, p.name, p.name_am
            ORDER BY order_count DESC LIMIT 5
        """, (today,))
        top_products = [dict(zip(['name','name_am','order_count'], r))
                        for r in cursor.fetchall()]

        cursor.execute("SELECT value FROM settings WHERE key='admin_email'")
        row = cursor.fetchone()
        admin_email = (row[0] if row else None) or os.environ.get('ADMIN_EMAIL', '')

        if not admin_email:
            return jsonify({'ok': False, 'msg': 'Admin email አልተዋቀረም። Settings ላይ "Admin Email" ያስቀምጡ።'})

        data = {
            'date': f"TEST — {today.strftime('%B %d, %Y')}",
            'new_orders': new_orders,
            'total_revenue': total_revenue,
            'pending_orders': pending_orders,
            'low_stock_products': low_stock,
            'ai_conversations': ai_conversations,
            'top_products': top_products,
        }
        subject = f"📊 [TEST] SEMIRA Daily Digest — {today.strftime('%b %d, %Y')}"
        ok = send_email(admin_email, subject, build_digest_html(data), build_digest_text(data))

        if ok:
            return jsonify({'ok': True, 'msg': f'✅ Test email sent to {admin_email}'})
        else:
            return jsonify({'ok': False, 'msg': '❌ Send failed — SMTP credentials ያረጋግጡ (Gmail App Password ይፈልጋሉ).'})

    except Exception as e:
        current_app.logger.error(f"send_test_digest error: {e}")
        return jsonify({'ok': False, 'msg': f'Error: {e}'})


@admin_bp.route('/ai-logs/clear', methods=['POST'])
@admin_required
def ai_logs_clear():
    """Delete all AI conversation logs."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM ai_conversations")
        conn.commit()
        flash('All AI conversation logs cleared.', 'success')
    except Exception as e:
        current_app.logger.error(f"ai_logs_clear error: {e}")
        flash('Error clearing logs.', 'error')
    return redirect(url_for('admin.ai_logs'))


@admin_bp.route('/settings/ai-key', methods=['POST'])
@admin_required
def save_ai_key():
    """Save or update GROQ_API_KEY in the settings table and live environment."""
    key = request.form.get('groq_api_key', '').strip()
    if not key:
        flash('No API key entered — existing key kept.', 'info')
        return redirect(url_for('admin.settings') + '#ai-settings')
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO settings (key, value) VALUES (%s, %s)
            ON CONFLICT(key) DO UPDATE SET value = %s
        """, ('groq_api_key', key, key))
        conn.commit()
        os.environ['GROQ_API_KEY'] = key
        flash('Groq API key saved! AI assistant is now using Groq LLM.', 'success')
    except Exception as e:
        current_app.logger.error(f"save_ai_key error: {e}")
        flash('Error saving API key.', 'error')
    return redirect(url_for('admin.settings') + '#ai-settings')


@admin_bp.route('/settings/remove-ai-key', methods=['POST'])
@admin_required
def remove_ai_key():
    """Remove GROQ_API_KEY from settings and environment."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM settings WHERE key = %s", ('groq_api_key',))
        conn.commit()
        os.environ.pop('GROQ_API_KEY', None)
        return jsonify({'success': True})
    except Exception as e:
        current_app.logger.error(f"remove_ai_key error: {e}")
        return jsonify({'success': False, 'error': str(e)})


@admin_bp.route('/settings/test-ai', methods=['POST'])
@admin_required
def test_ai_key():
    """Test if the configured GROQ_API_KEY works by sending a minimal request."""
    api_key = os.environ.get('GROQ_API_KEY', '').strip()
    if not api_key:
        return jsonify({'success': False, 'error': 'No GROQ_API_KEY configured.'})
    try:
        from groq import Groq
        client = Groq(api_key=api_key)
        completion = client.chat.completions.create(
            model='llama-3.3-70b-versatile',
            messages=[{'role': 'user', 'content': 'Say OK in one word.'}],
            max_tokens=5,
            temperature=0,
        )
        reply = completion.choices[0].message.content.strip()
        return jsonify({'success': True, 'model': 'llama-3.3-70b-versatile', 'sample': reply})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)[:200]})


@admin_bp.route('/clear-cache', methods=['GET', 'POST'])
@admin_required
def clear_cache():
    try:
        session.pop('cart', None)
        from flask import current_app
        current_app.jinja_env.cache = {}
        return jsonify({'success': True, 'message': 'Cache cleared successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/backup-database', methods=['GET', 'POST'])
@admin_required
def backup_database():
    try:
        return jsonify({
            'success': False,
            'error': 'This app uses PostgreSQL. Use pg_dump from the shell or Replit Database panel to back up.'
        }), 400
    except Exception as e:
        current_app.logger.error(f"Backup error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== NOTIFICATIONS ====================

@admin_bp.route('/send-notification', methods=['GET', 'POST'])
@admin_required
def send_notification():
    lang = get_lang()
    if request.method == 'POST':
        try:
            title = request.form.get('title', '').strip()
            body = request.form.get('body', '').strip()
            title_am = request.form.get('title_am', '').strip()
            title_ar = request.form.get('title_ar', '').strip()
            body_am = request.form.get('body_am', '').strip()
            body_ar = request.form.get('body_ar', '').strip()
            image_url = request.form.get('image_url', '').strip()
            link = request.form.get('link', '').strip()
            target = request.form.get('target', 'all')

            if not title or not body:
                flash('Please enter both title and message', 'error')
                return redirect(url_for('admin.send_notification'))

            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO notifications (
                    title, title_am, title_ar, body, body_am, body_ar,
                    image, link, target_audience, sent_at, created_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, %s)
            """, (title, title_am, title_ar, body, body_am, body_ar,
                  image_url, link, target, session.get('admin_username', 'admin')))
            conn.commit()

            # Fan-out to user_notifications table so customers see it in dashboard
            try:
                cursor.execute("SELECT id FROM users WHERE is_admin = 0 AND is_active = 1")
                user_rows = cursor.fetchall()
                for ur in (user_rows or []):
                    uid = ur[0]
                    notify_user(uid, title, body, type='info', link=link or '')
            except Exception as fan_err:
                current_app.logger.warning(f"⚠️ Notification fan-out partial failure: {fan_err}")

            flash('Notification sent to all customers!', 'success')
        except Exception as e:
            current_app.logger.error(f"Send notification error: {e}")
            flash('Error sending notification.', 'error')
        return redirect(url_for('admin.send_notification'))

    return render_template('admin/send_notification.html', lang=lang)


@admin_bp.route('/alerts')
@admin_required
def admin_alerts_page():
    """Admin alerts inbox page."""
    lang = get_lang()
    alerts = get_admin_alerts(limit=100)
    unread = get_admin_unread_count()
    return render_template('admin/alerts.html', alerts=alerts, unread=unread, lang=lang)


@admin_bp.route('/alerts/mark-read', methods=['POST'])
@admin_required
def admin_mark_alerts_read():
    alert_id = request.json.get('id') if request.is_json else None
    mark_admin_alerts_read(alert_id)
    return jsonify({'success': True})


# ==================== REVIEWS ====================

@admin_bp.route('/reviews')
@admin_required
def admin_reviews():
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT r.*, p.name_am as product_name, u.full_name as user_name
            FROM reviews r
            JOIN products p ON r.product_id = p.id
            JOIN users u ON r.user_id = u.id
            ORDER BY r.created_at DESC
        """)
        reviews_rows = cursor.fetchall()
        return render_template('admin/reviews/index.html',
                               reviews=[dict(r) for r in reviews_rows] if reviews_rows else [],
                               lang=lang)
    except Exception as e:
        current_app.logger.error(f"Admin reviews error: {e}")
        flash('Error loading reviews', 'error')
        return redirect(url_for('admin.dashboard'))


@admin_bp.route('/reviews/approve/<int:review_id>', methods=['POST'])
@admin_required
def approve_review(review_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("UPDATE reviews SET is_approved = 1 WHERE id = %s", (review_id,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Review approved'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/reviews/delete/<int:review_id>', methods=['DELETE', 'POST'])
@admin_required
def delete_review(review_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM reviews WHERE id = %s", (review_id,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Review deleted'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== TRANSLATIONS ====================

@admin_bp.route('/translations/status', methods=['GET'])
@admin_required
def translation_status():
    try:
        from utils.translation_cache import get_translation_stats, FALLBACK_TEXTS
        stats = get_translation_stats()
        return jsonify({
            'status': 'success',
            'cache': stats,
            'fallback_languages': list(FALLBACK_TEXTS.keys()),
            'supported_languages': ['am', 'en', 'ar']
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@admin_bp.route('/translations/clear', methods=['POST'])
@admin_required
def clear_translations():
    try:
        from utils.translation_cache import clear_translation_cache
        clear_translation_cache()
        return jsonify({'status': 'success', 'message': 'Translation cache cleared successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ==================== PASSWORD RESET TOKENS ====================

@admin_bp.route('/password-resets')
@admin_required
def password_resets():
    """View pending password reset tokens — useful when email is not configured."""
    lang = get_lang()
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT prt.id, prt.email, prt.token, prt.expires_at, prt.created_at,
                   u.full_name
            FROM password_reset_tokens prt
            LEFT JOIN users u ON u.email = prt.email
            WHERE prt.used = 0
              AND prt.expires_at > NOW()
            ORDER BY prt.created_at DESC
        """)
        tokens = [dict(r) for r in (cursor.fetchall() or [])]

        import datetime as _dt
        from flask import request as _req
        base = _req.host_url.rstrip('/')
        for t in tokens:
            t['reset_url'] = f"{base}/reset-password/{t['token']}"
            if t['expires_at']:
                remaining = t['expires_at'] - _dt.datetime.now()
                mins = int(remaining.total_seconds() / 60)
                t['expires_in'] = f"{mins} min" if mins > 0 else "expired"

        return render_template('admin/password_resets/index.html',
                               tokens=tokens, lang=lang)
    except Exception as e:
        current_app.logger.error(f"Password resets admin error: {e}")
        flash('Error loading reset tokens.', 'error')
        return redirect(url_for('admin.dashboard'))


@admin_bp.route('/password-resets/invalidate/<int:token_id>', methods=['POST'])
@admin_required
def invalidate_reset_token(token_id):
    """Manually invalidate a reset token."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("UPDATE password_reset_tokens SET used = 1 WHERE id = %s", (token_id,))
        conn.commit()
        flash('Reset token invalidated.', 'success')
    except Exception as e:
        current_app.logger.error(f"Invalidate token error: {e}")
        flash('Error invalidating token.', 'error')
    return redirect(url_for('admin.password_resets'))
